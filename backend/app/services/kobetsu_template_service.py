"""
Kobetsu Template Service - Template-based document generation.

This service generates documents by filling an Excel template (個別契約書TEXPERT2025.12Perfect.xlsm)
which contains 18 interconnected sheets for all dispatch document types.

Template Architecture:
- Control Panel: Cells AD1-AD5 in 個別契約書X sheet
- Formula System: XLOOKUP against TBKaishaInfo table
- Document Sheets: 個別契約書X, 通知書, DAICHO, 派遣元管理台帳, etc.

For web-based generation, we set both:
1. Control cells (AD1-AD5, dates)
2. Formula cells directly (to avoid TBKaishaInfo dependency)
"""
import logging
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core.config import settings

logger = logging.getLogger(__name__)


class KobetsuTemplateService:
    """
    Generates dispatch documents from Excel template.

    Template: 個別契約書TEXPERT2025.12Perfect.xlsm (18 sheets)

    Document Types:
    - 個別契約書 (Individual Contract)
    - 通知書 (Notification)
    - DAICHO (派遣先管理台帳)
    - 派遣元管理台帳 (Dispatch Origin Registry)
    - 就業条件明示書 (Employment Conditions)
    - 契約書 (Employment Contract)
    - 雇入れ時の待遇情報 (Treatment Info at Hiring)
    - タイムシート (Timesheet)
    - 就業状況報告書 (Employment Status Report)
    """

    # Template file name
    TEMPLATE_FILENAME = "個別契約書TEXPERT2025.12Perfect.xlsm"

    # Sheet names
    SHEET_KOBETSU = "個別契約書X"
    SHEET_TSUCHISHO = "通知書"
    SHEET_TSUCHISHO_BETSUBETSU = "通知書(別々)"
    SHEET_DAICHO = "DAICHO"
    SHEET_HAKENMOTO_DAICHO = "派遣元管理台帳"
    SHEET_SHUGYO_JOKEN = "就業条件明示書"
    SHEET_KEIYAKUSHO = "契約書"
    SHEET_YATOIRE_TAIGU = "雇入れ時の待遇情報"
    SHEET_TIMESHEET = " タイムシート"
    SHEET_SHUGYO_HOKOKU = "就業状況報告書（本社)"

    def __init__(self, template_path: Optional[Path] = None):
        """
        Initialize the service.

        Args:
            template_path: Optional custom template path
        """
        self.template_path = self._find_template(template_path)
        logger.info(f"Using template: {self.template_path}")

    def _find_template(self, custom_path: Optional[Path] = None) -> Path:
        """Find the template file in various locations."""
        if custom_path and custom_path.exists():
            return custom_path

        # Search locations in order of preference
        search_paths = [
            Path("/app/templates") / self.TEMPLATE_FILENAME,
            Path("templates") / self.TEMPLATE_FILENAME,
            Path("backend/templates") / self.TEMPLATE_FILENAME,
            Path(self.TEMPLATE_FILENAME),
            Path(__file__).parent.parent.parent.parent.parent / self.TEMPLATE_FILENAME,
        ]

        for path in search_paths:
            if path.exists():
                return path

        raise FileNotFoundError(
            f"Template '{self.TEMPLATE_FILENAME}' not found. "
            f"Searched: {[str(p) for p in search_paths]}"
        )

    # ========================================================================
    # PUBLIC API - Generate Documents
    # ========================================================================

    def generate_kobetsu_keiyakusho(
        self,
        contract,
        factory=None,
        employees: Optional[List] = None
    ) -> bytes:
        """
        Generate 個別契約書 (Individual Dispatch Contract).

        Args:
            contract: KobetsuKeiyakusho model
            factory: Factory model (optional, uses contract.factory if not provided)
            employees: List of employees (optional)

        Returns:
            bytes: .xlsm file content
        """
        if factory is None:
            factory = getattr(contract, 'factory', None)

        data = self._build_data_from_models(contract, factory, employees)
        return self._generate_document(data, target_sheet=self.SHEET_KOBETSU)

    def generate_tsuchisho(
        self,
        contract,
        factory=None,
        employees: Optional[List] = None
    ) -> bytes:
        """
        Generate 通知書 (Dispatch Notification).

        Args:
            contract: KobetsuKeiyakusho model
            factory: Factory model
            employees: List of employees to notify

        Returns:
            bytes: .xlsm file content
        """
        if factory is None:
            factory = getattr(contract, 'factory', None)

        data = self._build_data_from_models(contract, factory, employees)
        return self._generate_document(data, target_sheet=self.SHEET_TSUCHISHO)

    def generate_daicho(
        self,
        contract,
        factory=None,
        employee=None
    ) -> bytes:
        """
        Generate DAICHO (派遣先管理台帳 - Destination Registry).

        Args:
            contract: KobetsuKeiyakusho model
            factory: Factory model
            employee: Single employee for the registry

        Returns:
            bytes: .xlsm file content
        """
        if factory is None:
            factory = getattr(contract, 'factory', None)

        employees = [employee] if employee else None
        data = self._build_data_from_models(contract, factory, employees)
        return self._generate_document(data, target_sheet=self.SHEET_DAICHO)

    def generate_hakenmoto_daicho(
        self,
        contract,
        factory=None,
        employee=None
    ) -> bytes:
        """
        Generate 派遣元管理台帳 (Dispatch Origin Registry).

        Args:
            contract: KobetsuKeiyakusho model
            factory: Factory model
            employee: Single employee for the registry

        Returns:
            bytes: .xlsm file content
        """
        if factory is None:
            factory = getattr(contract, 'factory', None)

        employees = [employee] if employee else None
        data = self._build_data_from_models(contract, factory, employees)
        return self._generate_document(data, target_sheet=self.SHEET_HAKENMOTO_DAICHO)

    def generate_shugyo_joken(
        self,
        contract,
        factory=None,
        employee=None
    ) -> bytes:
        """
        Generate 就業条件明示書 (Employment Conditions Statement).

        Args:
            contract: KobetsuKeiyakusho model
            factory: Factory model
            employee: Single employee

        Returns:
            bytes: .xlsm file content
        """
        if factory is None:
            factory = getattr(contract, 'factory', None)

        employees = [employee] if employee else None
        data = self._build_data_from_models(contract, factory, employees)
        return self._generate_document(data, target_sheet=self.SHEET_SHUGYO_JOKEN)

    def generate_keiyakusho(
        self,
        contract,
        factory=None,
        employee=None
    ) -> bytes:
        """
        Generate 契約書 (Employment Contract).

        Args:
            contract: KobetsuKeiyakusho model
            factory: Factory model
            employee: Single employee

        Returns:
            bytes: .xlsm file content
        """
        if factory is None:
            factory = getattr(contract, 'factory', None)

        employees = [employee] if employee else None
        data = self._build_data_from_models(contract, factory, employees)
        return self._generate_document(data, target_sheet=self.SHEET_KEIYAKUSHO)

    def generate_from_data(self, data: Dict[str, Any], document_type: str = "kobetsu") -> bytes:
        """
        Generate document from raw data dictionary.

        Args:
            data: Dictionary with all contract data fields
            document_type: One of 'kobetsu', 'tsuchisho', 'daicho', 'hakenmoto_daicho',
                          'shugyo_joken', 'keiyakusho', 'yatoire_taigu'

        Returns:
            bytes: .xlsm file content
        """
        sheet_map = {
            "kobetsu": self.SHEET_KOBETSU,
            "tsuchisho": self.SHEET_TSUCHISHO,
            "daicho": self.SHEET_DAICHO,
            "hakenmoto_daicho": self.SHEET_HAKENMOTO_DAICHO,
            "shugyo_joken": self.SHEET_SHUGYO_JOKEN,
            "keiyakusho": self.SHEET_KEIYAKUSHO,
            "yatoire_taigu": self.SHEET_YATOIRE_TAIGU,
        }

        target_sheet = sheet_map.get(document_type, self.SHEET_KOBETSU)
        return self._generate_document(data, target_sheet=target_sheet)

    # ========================================================================
    # INTERNAL - Document Generation
    # ========================================================================

    def _generate_document(self, data: Dict[str, Any], target_sheet: str) -> bytes:
        """
        Generate document by filling template cells.

        Args:
            data: Data dictionary with all fields
            target_sheet: Target sheet name to activate for printing

        Returns:
            bytes: .xlsm file content
        """
        # Load template with VBA preserved
        wb = load_workbook(str(self.template_path), keep_vba=True)

        # Fill the control cells in 個別契約書X
        if self.SHEET_KOBETSU in wb.sheetnames:
            ws_control = wb[self.SHEET_KOBETSU]
            self._fill_control_cells(ws_control, data)
            self._fill_kobetsu_document_cells(ws_control, data)

        # Fill sheet-specific cells based on target
        if target_sheet in wb.sheetnames:
            ws_target = wb[target_sheet]

            if target_sheet == self.SHEET_TSUCHISHO:
                self._fill_tsuchisho_cells(ws_target, data)
            elif target_sheet == self.SHEET_DAICHO:
                self._fill_daicho_cells(ws_target, data)
            elif target_sheet == self.SHEET_HAKENMOTO_DAICHO:
                self._fill_hakenmoto_daicho_cells(ws_target, data)
            elif target_sheet == self.SHEET_SHUGYO_JOKEN:
                self._fill_shugyo_joken_cells(ws_target, data)
            elif target_sheet == self.SHEET_KEIYAKUSHO:
                self._fill_keiyakusho_cells(ws_target, data)

            # Set as active sheet
            wb.active = ws_target

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    # ========================================================================
    # CELL FILLING METHODS
    # ========================================================================

    def _fill_control_cells(self, ws: Worksheet, data: Dict[str, Any]) -> None:
        """
        Fill the control panel cells (AD1-AD8) in 個別契約書X.

        These cells drive the formulas in all document sheets.
        """
        # Main control cells
        self._set_cell(ws, 'AD1', data.get('company_name', data.get('派遣先', '')))
        self._set_cell(ws, 'AD2', data.get('factory_name', data.get('工場名', '')))
        self._set_cell(ws, 'AD3', data.get('department', data.get('配属先', '')))
        self._set_cell(ws, 'AD4', data.get('line', data.get('ライン', '')))
        self._set_cell(ws, 'AD5', data.get('hourly_rate', data.get('時給単価', '')))

        # Date controls
        start_date = data.get('dispatch_start_date')
        end_date = data.get('dispatch_end_date')

        if start_date:
            self._set_cell(ws, 'H16', self._to_excel_date(start_date))
        if end_date:
            self._set_cell(ws, 'P16', self._to_excel_date(end_date))

    def _fill_kobetsu_document_cells(self, ws: Worksheet, data: Dict[str, Any]) -> None:
        """
        Fill the document cells in 個別契約書X sheet.

        We set these directly to avoid formula dependencies on TBKaishaInfo.
        """
        # ================================================================
        # 派遣先 (Client) Section - Rows 4-9
        # ================================================================

        # Row 4: 派遣先事業所
        self._set_cell(ws, 'I4', data.get('client_company_name', data.get('company_name', '')))
        self._set_cell(ws, 'Q4', data.get('client_address', ''))
        self._set_cell(ws, 'Y4', data.get('client_tel', ''))

        # Row 5: 就業場所
        self._set_cell(ws, 'I5', data.get('worksite_name', ''))
        self._set_cell(ws, 'Q5', data.get('worksite_address', ''))
        self._set_cell(ws, 'Y5', data.get('worksite_tel', ''))

        # Row 6: 組織単位 & 抵触日
        self._set_cell(ws, 'H6', data.get('organizational_unit', ''))
        self._set_cell(ws, 'Q6', self._format_japanese_date(data.get('conflict_date')))

        # Row 7: 指揮命令者
        self._set_cell(ws, 'J7', data.get('supervisor_dept', ''))
        self._set_cell(ws, 'Q7', data.get('supervisor_position', ''))
        self._set_cell(ws, 'S7', data.get('supervisor_name', ''))
        self._set_cell(ws, 'Y7', data.get('supervisor_tel', ''))

        # Row 8: 製造業務専門派遣先責任者
        self._set_cell(ws, 'J8', data.get('saki_manager_dept', ''))
        self._set_cell(ws, 'Q8', data.get('saki_manager_position', ''))
        self._set_cell(ws, 'S8', data.get('saki_manager_name', ''))
        self._set_cell(ws, 'Y8', data.get('saki_manager_tel', ''))

        # Row 9: 派遣先苦情処理担当者
        self._set_cell(ws, 'J9', data.get('saki_complaint_dept', ''))
        self._set_cell(ws, 'Q9', data.get('saki_complaint_position', ''))
        self._set_cell(ws, 'S9', data.get('saki_complaint_name', ''))
        self._set_cell(ws, 'Y9', data.get('saki_complaint_tel', ''))

        # ================================================================
        # 派遣元 (Agency - UNS) Section - Rows 10-11
        # ================================================================

        # Row 10: 製造業務専門派遣元責任者
        self._set_cell(ws, 'J10', data.get('moto_manager_dept', settings.DISPATCH_RESPONSIBLE_DEPARTMENT))
        self._set_cell(ws, 'Q10', data.get('moto_manager_position', settings.DISPATCH_RESPONSIBLE_POSITION))
        self._set_cell(ws, 'S10', data.get('moto_manager_name', settings.DISPATCH_RESPONSIBLE_NAME))
        self._set_cell(ws, 'Y10', data.get('moto_manager_tel', settings.DISPATCH_RESPONSIBLE_PHONE))

        # Row 11: 派遣元苦情処理担当者
        self._set_cell(ws, 'J11', data.get('moto_complaint_dept', settings.DISPATCH_COMPLAINT_DEPARTMENT))
        self._set_cell(ws, 'Q11', data.get('moto_complaint_position', settings.DISPATCH_COMPLAINT_POSITION))
        self._set_cell(ws, 'S11', data.get('moto_complaint_name', settings.DISPATCH_COMPLAINT_NAME))
        self._set_cell(ws, 'Y11', data.get('moto_complaint_tel', settings.DISPATCH_COMPLAINT_PHONE))

        # ================================================================
        # 派遣内容 (Dispatch Content) Section - Rows 12-28
        # ================================================================

        # Row 15: 業務内容
        self._set_cell(ws, 'H15', data.get('work_content', data.get('job_description', '')))

        # Row 16: 派遣期間 & 人数
        self._set_cell(ws, 'H16', self._format_japanese_date(data.get('dispatch_start_date')))
        self._set_cell(ws, 'P16', self._format_japanese_date(data.get('dispatch_end_date')))
        self._set_cell(ws, 'Y16', data.get('number_of_workers', data.get('headcount', 1)))

        # Row 17: 就業日
        self._set_cell(ws, 'H17', data.get('work_days_text', self._format_work_days(data.get('work_days'))))

        # Row 18: 就業時間
        work_time_text = data.get('work_time_text')
        if not work_time_text:
            start = self._format_time(data.get('work_start_time'))
            end = self._format_time(data.get('work_end_time'))
            if start and end:
                work_time_text = f"昼勤：{start}～{end}"
        self._set_cell(ws, 'H18', work_time_text)

        # Row 22: 休憩時間
        self._set_cell(ws, 'H22', data.get('break_time_text', ''))

        # Row 26: 時間外労働
        self._set_cell(ws, 'H26', data.get('overtime_text', ''))

        # Row 27: 休日労働
        self._set_cell(ws, 'H27', data.get('holiday_work_text', ''))

        # ================================================================
        # 派遣料金 (Rates) Section - Rows 29-33
        # ================================================================

        # Row 29: 基本料金
        base_rate = data.get('hourly_rate', data.get('rate_base'))
        if base_rate:
            self._set_cell(ws, 'K29', base_rate)

            # Calculate derived rates if not provided
            overtime_rate = data.get('overtime_rate')
            if not overtime_rate:
                overtime_rate = float(base_rate) * 1.25
            self._set_cell(ws, 'Q29', overtime_rate)

            holiday_rate = data.get('holiday_rate')
            if not holiday_rate:
                holiday_rate = float(base_rate) * 1.35
            self._set_cell(ws, 'W29', holiday_rate)

        # Row 32-33: 支払条件
        self._set_cell(ws, 'K32', data.get('cutoff_day', ''))
        self._set_cell(ws, 'Q32', data.get('payment_day', ''))

        # ================================================================
        # 署名 (Signature) Section - Rows 59-65
        # ================================================================

        # Row 59: 契約締結日
        self._set_cell(ws, 'A59', self._to_excel_date(data.get('contract_date')))

        # Row 61: 派遣元住所
        self._set_cell(ws, 'O61', data.get('agency_address', settings.COMPANY_ADDRESS))

        # Row 62: 派遣元会社名
        self._set_cell(ws, 'O62', data.get('agency_company_name', settings.COMPANY_NAME))

        # Row 63: 代表者
        rep_title = data.get('representative_title', settings.DISPATCH_MANAGER_POSITION)
        rep_name = data.get('representative_name', settings.DISPATCH_MANAGER_NAME)
        self._set_cell(ws, 'O63', f"{rep_title}　{rep_name}")

        # Row 64: 許可番号
        self._set_cell(ws, 'O64', data.get('license_number', settings.COMPANY_LICENSE_NUMBER))

    def _fill_tsuchisho_cells(self, ws: Worksheet, data: Dict[str, Any]) -> None:
        """Fill cells for 通知書 (Notification) sheet."""
        # Employee list starts at row 9
        employees = data.get('employees', [])
        for i, emp in enumerate(employees[:58]):  # Max 58 employees per sheet
            row = 9 + i
            self._set_cell(ws, f'I{row}', emp.get('name', emp.get('full_name_kanji', '')))
            self._set_cell(ws, f'J{row}', emp.get('gender', ''))

    def _fill_daicho_cells(self, ws: Worksheet, data: Dict[str, Any]) -> None:
        """Fill cells for DAICHO (派遣先管理台帳) sheet."""
        # Employee info
        employees = data.get('employees', [])
        if employees:
            emp = employees[0]
            self._set_cell(ws, 'C11', emp.get('name', emp.get('full_name_kanji', '')))

    def _fill_hakenmoto_daicho_cells(self, ws: Worksheet, data: Dict[str, Any]) -> None:
        """Fill cells for 派遣元管理台帳 sheet."""
        # Employee info
        employees = data.get('employees', [])
        if employees:
            emp = employees[0]
            self._set_cell(ws, 'G6', emp.get('name', emp.get('full_name_kanji', '')))

        # Company info
        self._set_cell(ws, 'B9', data.get('client_company_name', ''))
        self._set_cell(ws, 'J9', data.get('worksite_name', ''))

    def _fill_shugyo_joken_cells(self, ws: Worksheet, data: Dict[str, Any]) -> None:
        """Fill cells for 就業条件明示書 sheet."""
        employees = data.get('employees', [])
        if employees:
            emp = employees[0]
            self._set_cell(ws, 'B6', emp.get('name', emp.get('full_name_kanji', '')))

    def _fill_keiyakusho_cells(self, ws: Worksheet, data: Dict[str, Any]) -> None:
        """Fill cells for 契約書 (Employment Contract) sheet."""
        employees = data.get('employees', [])
        if employees:
            emp = employees[0]
            self._set_cell(ws, 'C4', emp.get('full_name_kana', emp.get('katakana_name', '')))
            self._set_cell(ws, 'C5', emp.get('name', emp.get('full_name_kanji', '')))
            self._set_cell(ws, 'C6', emp.get('address', ''))
            self._set_cell(ws, 'C7', self._format_japanese_date(emp.get('date_of_birth')))
            self._set_cell(ws, 'C8', emp.get('gender', ''))

        # Contract dates
        self._set_cell(ws, 'C10', self._to_excel_date(data.get('dispatch_start_date')))
        self._set_cell(ws, 'G10', self._to_excel_date(data.get('dispatch_end_date')))

    # ========================================================================
    # HELPER METHODS
    # ========================================================================

    def _set_cell(self, ws: Worksheet, cell_ref: str, value: Any) -> None:
        """Set a cell value, preserving existing formatting."""
        if value is None or value == '':
            return

        try:
            cell = ws[cell_ref]
            cell.value = value
        except Exception as e:
            logger.warning(f"Failed to set cell {cell_ref}: {e}")

    def _to_excel_date(self, d: Optional[Union[date, datetime, str]]) -> Optional[Union[date, int]]:
        """Convert date to Excel serial date number or date object."""
        if d is None:
            return None

        if isinstance(d, str):
            try:
                d = datetime.fromisoformat(d.replace('Z', '+00:00')).date()
            except ValueError:
                return None

        if isinstance(d, datetime):
            d = d.date()

        return d

    def _format_japanese_date(self, d: Optional[Union[date, datetime, str]]) -> str:
        """Format date in Japanese format (YYYY年MM月DD日)."""
        if d is None:
            return ""

        if isinstance(d, str):
            try:
                d = datetime.fromisoformat(d.replace('Z', '+00:00')).date()
            except ValueError:
                return d

        if isinstance(d, datetime):
            d = d.date()

        if isinstance(d, date):
            return f"{d.year}年{d.month}月{d.day}日"

        return str(d)

    def _format_time(self, t: Optional[Union[time, datetime, str]]) -> str:
        """Format time in Japanese format (HH時MM分)."""
        if t is None:
            return ""

        if isinstance(t, str):
            if '時' in t:
                return t
            try:
                parts = t.split(':')
                if len(parts) >= 2:
                    return f"{int(parts[0])}時{int(parts[1]):02d}分"
            except (ValueError, IndexError):
                return t

        if isinstance(t, datetime):
            t = t.time()

        if isinstance(t, time):
            return f"{t.hour}時{t.minute:02d}分"

        return str(t)

    def _format_work_days(self, work_days: Optional[list]) -> str:
        """Format work days list to Japanese text."""
        if not work_days:
            return ""

        if isinstance(work_days, str):
            return work_days

        if isinstance(work_days, list):
            if len(work_days) >= 5 and set(work_days) >= {'月', '火', '水', '木', '金'}:
                return "月〜金（シフトに準ずる）"
            return '・'.join(work_days)

        return str(work_days)

    def _build_data_from_models(
        self,
        contract,
        factory=None,
        employees: Optional[List] = None
    ) -> Dict[str, Any]:
        """
        Build data dictionary from SQLAlchemy models.

        Args:
            contract: KobetsuKeiyakusho model
            factory: Factory model
            employees: List of Employee or KobetsuEmployee models

        Returns:
            Dictionary with all data fields
        """
        data = {}

        # Factory/Company info
        if factory:
            data.update({
                'company_name': factory.company_name,
                'factory_name': factory.plant_name or '',
                'department': getattr(factory, 'organizational_unit', ''),
                'line': getattr(factory, 'line', ''),
                'client_company_name': factory.company_name,
                'client_address': factory.company_address or '',
                'client_tel': getattr(factory, 'company_phone', ''),
                'worksite_name': f"{factory.company_name} {factory.plant_name or ''}".strip(),
                'worksite_address': factory.plant_address or factory.company_address or '',
                'worksite_tel': getattr(factory, 'plant_phone', ''),
                'conflict_date': getattr(factory, 'conflict_date', None),
            })

            # Supervisor info from factory
            if hasattr(factory, 'supervisor_name'):
                data['supervisor_name'] = factory.supervisor_name
            if hasattr(factory, 'supervisor_department'):
                data['supervisor_dept'] = factory.supervisor_department
            if hasattr(factory, 'supervisor_position'):
                data['supervisor_position'] = factory.supervisor_position

        # Contract info
        if contract:
            data.update({
                'dispatch_start_date': contract.dispatch_start_date,
                'dispatch_end_date': contract.dispatch_end_date,
                'contract_date': contract.contract_date,
                'work_content': contract.work_content,
                'job_description': contract.work_content,
                'number_of_workers': contract.number_of_workers,
                'headcount': contract.number_of_workers,
                'work_days': contract.work_days,
                'work_start_time': contract.work_start_time,
                'work_end_time': contract.work_end_time,
                'break_time_minutes': contract.break_time_minutes,
                'hourly_rate': float(contract.hourly_rate) if contract.hourly_rate else None,
                'overtime_rate': float(contract.overtime_rate) if contract.overtime_rate else None,
                'holiday_rate': float(contract.holiday_rate) if contract.holiday_rate else None,
                'organizational_unit': contract.organizational_unit,
                'supervisor_name': contract.supervisor_name,
                'supervisor_dept': contract.supervisor_department,
                'supervisor_position': contract.supervisor_position,
                'safety_measures': contract.safety_measures,
            })

            # Manager info from contract
            if contract.haken_saki_manager:
                data.update({
                    'saki_manager_dept': contract.haken_saki_manager.get('department', ''),
                    'saki_manager_position': contract.haken_saki_manager.get('position', ''),
                    'saki_manager_name': contract.haken_saki_manager.get('name', ''),
                    'saki_manager_tel': contract.haken_saki_manager.get('phone', ''),
                })

            if contract.haken_saki_complaint_contact:
                data.update({
                    'saki_complaint_dept': contract.haken_saki_complaint_contact.get('department', ''),
                    'saki_complaint_position': contract.haken_saki_complaint_contact.get('position', ''),
                    'saki_complaint_name': contract.haken_saki_complaint_contact.get('name', ''),
                    'saki_complaint_tel': contract.haken_saki_complaint_contact.get('phone', ''),
                })

        # Employee info
        if employees:
            emp_list = []
            for emp in employees:
                # Handle both Employee and KobetsuEmployee models
                if hasattr(emp, 'employee'):
                    actual_emp = emp.employee
                else:
                    actual_emp = emp

                emp_list.append({
                    'name': getattr(actual_emp, 'full_name_kanji', ''),
                    'full_name_kanji': getattr(actual_emp, 'full_name_kanji', ''),
                    'full_name_kana': getattr(actual_emp, 'full_name_kana', ''),
                    'katakana_name': getattr(actual_emp, 'full_name_kana', ''),
                    'gender': getattr(actual_emp, 'gender', ''),
                    'date_of_birth': getattr(actual_emp, 'date_of_birth', None),
                    'address': getattr(actual_emp, 'address', ''),
                    'employee_number': getattr(actual_emp, 'employee_number', ''),
                })

            data['employees'] = emp_list

        # Agency defaults
        data.update({
            'agency_company_name': settings.COMPANY_NAME,
            'agency_address': settings.COMPANY_ADDRESS,
            'moto_manager_dept': settings.DISPATCH_RESPONSIBLE_DEPARTMENT,
            'moto_manager_position': settings.DISPATCH_RESPONSIBLE_POSITION,
            'moto_manager_name': settings.DISPATCH_RESPONSIBLE_NAME,
            'moto_manager_tel': settings.DISPATCH_RESPONSIBLE_PHONE,
            'moto_complaint_dept': settings.DISPATCH_COMPLAINT_DEPARTMENT,
            'moto_complaint_position': settings.DISPATCH_COMPLAINT_POSITION,
            'moto_complaint_name': settings.DISPATCH_COMPLAINT_NAME,
            'moto_complaint_tel': settings.DISPATCH_COMPLAINT_PHONE,
            'representative_title': settings.DISPATCH_MANAGER_POSITION,
            'representative_name': settings.DISPATCH_MANAGER_NAME,
            'license_number': settings.COMPANY_LICENSE_NUMBER,
        })

        return data


# ============================================================================
# CONVENIENCE FUNCTIONS
# ============================================================================

def generate_kobetsu_from_template(contract, factory=None, employees=None) -> bytes:
    """Generate 個別契約書 from template."""
    service = KobetsuTemplateService()
    return service.generate_kobetsu_keiyakusho(contract, factory, employees)


def generate_tsuchisho_from_template(contract, factory=None, employees=None) -> bytes:
    """Generate 通知書 from template."""
    service = KobetsuTemplateService()
    return service.generate_tsuchisho(contract, factory, employees)


def generate_daicho_from_template(contract, factory=None, employee=None) -> bytes:
    """Generate DAICHO from template."""
    service = KobetsuTemplateService()
    return service.generate_daicho(contract, factory, employee)
