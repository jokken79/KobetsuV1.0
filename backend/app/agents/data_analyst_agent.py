from typing import Any, Dict, List, Tuple
import pandas as pd
from io import BytesIO
from datetime import datetime
from .base_agent import BaseAgent

class DataAnalystAgent(BaseAgent):
    """
    Agent responsible for analyzing and parsing data files (Excel/CSV).
    Currently implements heuristic logic for Employee and Factory imports.
    """

    def __init__(self):
        super().__init__(name="DataAnalystAgent", role="Data Processing Specialist")

    def process(self, input_data: Any) -> Dict[str, Any]:
        """
        Process generic input. For this agent, we expect specific methods to be called
        like 'preview_employees' or 'preview_factories'.
        
        This generic method could be used as a router in the future.
        """
        return {"message": "Use specific methods: preview_employees, preview_factories"}

    def preview_employees(self, content: bytes, filename: str) -> Dict[str, Any]:
        """
        Analyze and parse Employee Excel file.
        Wraps the logic previously found in ImportService.
        """
        self.log_activity(f"Analyzing employee file: {filename}")
        
        try:
            # 1. Detect format and load data
            df, is_dbgenzai, sheet_name = self._load_employee_excel(content)
            self.log_activity(f"Detected format: {'DBGenzaiX' if is_dbgenzai else 'Simple Template'} (Sheet: {sheet_name})")

            # 2. Map columns
            column_mapping = self._get_employee_column_mapping(is_dbgenzai)
            df = df.rename(columns=column_mapping)
            
            # 3. Process rows
            preview_data = []
            errors = []
            
            for idx, row in df.iterrows():
                row_num = idx + 2
                emp_data = self._clean_row_data(row)
                
                # Apply transformations
                if is_dbgenzai:
                    emp_data = self._transform_dbgenzai_data(emp_data)

                # Validate
                row_errors = self._validate_employee_data(row_num, emp_data, is_dbgenzai)
                if row_errors:
                    errors.extend(row_errors)

                preview_item = {
                    "row": row_num,
                    "employee_number": str(emp_data.get("employee_number", "")) if emp_data.get("employee_number") else "",
                    "full_name_kanji": emp_data.get("full_name_kanji", ""),
                    "full_name_kana": emp_data.get("full_name_kana", ""),
                    "company_name": emp_data.get("company_name", ""),
                    "hourly_rate": emp_data.get("hourly_rate"),
                    "hire_date": str(emp_data.get("hire_date", "")) if emp_data.get("hire_date") else "",
                    "is_valid": len(row_errors) == 0,
                    "errors": row_errors,
                    "_raw": emp_data
                }
                preview_data.append(preview_item)

            valid_count = len([p for p in preview_data if p["is_valid"]])
            
            return {
                "success": valid_count > 0,
                "total_rows": len(preview_data),
                "preview_data": preview_data,
                "errors": [], # Top level errors
                "message": f"Parsed {len(preview_data)} rows from '{sheet_name}'. Valid: {valid_count}."
            }

        except Exception as e:
            self.log_activity(f"Error parsing file: {str(e)}")
            return {
                "success": False,
                "total_rows": 0,
                "preview_data": [],
                "errors": [{"row": 0, "field": "file", "message": f"Error: {str(e)}"}],
                "message": f"Failed to parse file: {str(e)}"
            }

    def _load_employee_excel(self, content: bytes) -> Tuple[pd.DataFrame, bool, str]:
        """Load Excel and detect format."""
        from openpyxl import load_workbook
        
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        available_sheets = wb.sheetnames
        wb.close()

        # Sheet detection logic
        target_sheet = None
        for sheet in ['DBGenzaiX', 'DBGenzai', 'DB現在']:
            if sheet in available_sheets:
                target_sheet = sheet
                break
        if not target_sheet:
            target_sheet = available_sheets[0]

        df = pd.read_excel(BytesIO(content), sheet_name=target_sheet, engine='openpyxl')
        
        # Format detection
        is_dbgenzai = '社員№' in df.columns and '配属ライン' in df.columns
        
        return df, is_dbgenzai, target_sheet

    def _get_employee_column_mapping(self, is_dbgenzai: bool) -> Dict[str, str]:
        """Return column mapping based on detected format."""
        if is_dbgenzai:
            return {
                '現在': 'status_raw',
                '社員№': 'employee_number',
                '派遣先': 'company_name',
                '配属先': 'department',
                '配属ライン': 'line_name',
                '仕事内容': 'position',
                '氏名': 'full_name_kanji',
                'カナ': 'full_name_kana',
                '性別': 'gender',
                '国籍': 'nationality',
                '生年月日': 'date_of_birth',
                '時給': 'hourly_rate',
                '請求単価': 'billing_rate',
                'ビザ期限': 'visa_expiry_date',
                'ビザ種類': 'visa_type',
                '〒': 'postal_code',
                '住所': 'address',
                'アパート': 'apartment_name',
                'ｱﾊﾟｰﾄ': 'apartment_name',
                '入社日': 'hire_date',
                '退社日': 'termination_date',
                '社保加入': 'insurance_status',
                '備考': 'notes',
                '免許種類': 'drivers_license',
                '免許期限': 'drivers_license_expiry',
                '日本語検定': 'qualifications',
            }
        else:
            return {
                '社員№': 'employee_number',
                '社員番号': 'employee_number',
                '氏名': 'full_name_kanji',
                'カナ': 'full_name_kana',
                'ローマ字': 'full_name_romaji',
                '性別': 'gender',
                '生年月日': 'date_of_birth',
                '国籍': 'nationality',
                '住所': 'address',
                '電話番号': 'phone',
                '携帯電話': 'mobile',
                '入社日': 'hire_date',
                '退社日': 'termination_date',
                '派遣先': 'company_name',
                '工場': 'plant_name',
                '配属先': 'department',
                'ライン': 'line_name',
                '時給': 'hourly_rate',
                '請求単価': 'billing_rate',
                '在留資格': 'visa_type',
                'ビザ期限': 'visa_expiry_date',
                '在留カード番号': 'zairyu_card_number',
                '雇用保険': 'has_employment_insurance',
                '健康保険': 'has_health_insurance',
                '厚生年金': 'has_pension_insurance',
                '備考': 'notes',
            }

    def _clean_row_data(self, row: pd.Series) -> Dict[str, Any]:
        """Clean individual row data."""
        data = row.to_dict()
        return {k: (None if pd.isna(v) else v) for k, v in data.items()}

    def _transform_dbgenzai_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Apply DBGenzaiX specific transformations."""
        # Status
        if 'status_raw' in data:
            status_raw = data.get('status_raw')
            if status_raw:
                status_str = str(status_raw).strip().lower()
                if '退社' in status_str:
                    data['status'] = 'resigned'
                elif '休職' in status_str:
                    data['status'] = 'on_leave'
                else:
                    data['status'] = 'active'
            else:
                data['status'] = 'active'

        # Insurance
        if 'insurance_status' in data:
            insurance = data.get('insurance_status')
            if insurance and str(insurance).upper() == 'OK':
                data['has_health_insurance'] = True
                data['has_pension_insurance'] = True
                data['has_employment_insurance'] = True
        
        return data

    def _validate_employee_data(self, row_num: int, data: Dict[str, Any], is_dbgenzai: bool) -> List[str]:
        """Validate critical fields."""
        errors = []
        
        # Employee Number
        emp_num = data.get("employee_number")
        if not emp_num or (isinstance(emp_num, str) and not emp_num.strip()):
            errors.append(f"Row {row_num}: 社員番号は必須です")

        # Name
        if not data.get("full_name_kanji"):
            errors.append(f"Row {row_num}: 氏名は必須です")
            
        # Kana
        if not is_dbgenzai and not data.get("full_name_kana"):
            errors.append(f"Row {row_num}: カナは必須です")

        return errors
