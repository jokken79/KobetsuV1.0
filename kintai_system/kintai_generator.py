#!/usr/bin/env python3
"""
UNS 勤怠管理システム - Kintai Management System
=====================================
Sistema de registro de asistencia y generación de nómina para派遣社員

Estructura:
- 従業員マスタ: Datos de empleados
- 派遣先マスタ: Datos de fábricas/clientes  
- 勤怠表: Registro diario de horas
- 給与計算: Cálculo de salarios con割増
- 給与明細: Recibo de nómina individual
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
import calendar

# ===============================================
# CONFIGURACIÓN UNS
# ===============================================

UNS_CONFIG = {
    "company_name": "ユニバーサル企画株式会社",
    "permit_number": "派23-303669",
    "address": "愛知県名古屋市港区名港二丁目6-30",
    "phone": "052-938-8840",
    "bank_account": "愛知銀行 当知支店 普通2075479"
}

# 控除率 (Deduction rates)
DEDUCTION_RATES = {
    "health_insurance": Decimal("0.05"),      # 健康保険 5%
    "pension": Decimal("0.0915"),              # 厚生年金 9.15%
    "employment_insurance": Decimal("0.006"), # 雇用保険 0.6%
}

# 割増率 (Premium rates)
PREMIUM_RATES = {
    "overtime": Decimal("1.25"),      # 時間外 25%増
    "late_night": Decimal("1.25"),    # 深夜 25%増
    "holiday": Decimal("1.35"),       # 休日 35%増
    "overtime_60h": Decimal("1.50"),  # 60h超過 50%増
}

# スタイル定義
STYLES = {
    "header": {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill("solid", fgColor="2B5797"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    },
    "subheader": {
        "font": Font(bold=True, size=10),
        "fill": PatternFill("solid", fgColor="B4C6E7"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    },
    "data": {
        "font": Font(size=10),
        "alignment": Alignment(horizontal="left", vertical="center"),
    },
    "number": {
        "font": Font(size=10),
        "alignment": Alignment(horizontal="right", vertical="center"),
    },
    "currency": {
        "font": Font(size=10),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "number_format": "¥#,##0",
    },
    "percent": {
        "font": Font(size=10),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "number_format": "0.0%",
    },
    "time": {
        "font": Font(size=10),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "number_format": "h:mm",
    },
    "total": {
        "font": Font(bold=True, size=11),
        "fill": PatternFill("solid", fgColor="FFF2CC"),
        "alignment": Alignment(horizontal="right", vertical="center"),
    },
}

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def apply_style(cell, style_name):
    """Apply predefined style to cell"""
    style = STYLES.get(style_name, STYLES["data"])
    if "font" in style:
        cell.font = style["font"]
    if "fill" in style:
        cell.fill = style["fill"]
    if "alignment" in style:
        cell.alignment = style["alignment"]
    if "number_format" in style:
        cell.number_format = style["number_format"]
    cell.border = THIN_BORDER


def create_employee_master(wb):
    """Create 従業員マスタ sheet"""
    ws = wb.create_sheet("従業員マスタ")
    
    headers = [
        "社員番号", "氏名", "氏名カナ", "生年月日", "性別",
        "在留カード番号", "在留期限", "国籍", "住所", "電話番号",
        "派遣先", "工場名", "部署", "ライン", "入社日",
        "時給", "交通費", "社宅", "社宅家賃", "扶養人数"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_style(cell, "header")
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Sample data (Vietnamese workers at Kato Mokuzai)
    sample_employees = [
        ["E001", "グエン・ヴァン・ミン", "グエン ヴァン ミン", "1995-05-15", "男",
         "AB12345678CD", "2026-03-15", "ベトナム", "愛知県名古屋市港区当知1-1-1", "090-1234-5678",
         "加藤木材工業株式会社", "本社工場", "生産1部", "1課", "2023-04-01",
         1700, 5000, "有", 25000, 0],
        ["E002", "チャン・ティ・ラン", "チャン ティ ラン", "1998-08-22", "女",
         "EF98765432GH", "2025-12-20", "ベトナム", "愛知県名古屋市港区当知1-1-2", "090-8765-4321",
         "加藤木材工業株式会社", "本社工場", "生産1部", "2課", "2023-06-15",
         1700, 5000, "有", 25000, 0],
        ["E003", "レ・ヴァン・ドゥック", "レ ヴァン ドゥック", "1993-02-10", "男",
         "IJ11223344KL", "2026-06-30", "ベトナム", "愛知県名古屋市港区当知1-1-3", "090-1111-2222",
         "高雄工業株式会社", "岡山工場", "製作課", "CVJ短軸旋盤係", "2022-10-01",
         1650, 8000, "有", 28000, 1],
    ]
    
    for row_idx, emp in enumerate(sample_employees, 2):
        for col_idx, value in enumerate(emp, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in [16, 17, 19]:  # Currency columns
                apply_style(cell, "currency")
            else:
                apply_style(cell, "data")
    
    # Freeze header row
    ws.freeze_panes = "A2"
    return ws


def create_factory_master(wb):
    """Create 派遣先マスタ sheet"""
    ws = wb.create_sheet("派遣先マスタ")
    
    headers = [
        "派遣先ID", "派遣先名", "工場名", "住所", "電話番号",
        "担当者部署", "担当者名", "締め日", "支払日",
        "昼勤開始", "昼勤終了", "夜勤開始", "夜勤終了",
        "休憩時間(分)", "実働時間", "丸め単位(分)",
        "基本時給", "残業率", "深夜率", "休日率"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_style(cell, "header")
        ws.column_dimensions[get_column_letter(col)].width = 14
    
    factories = [
        ["F001", "加藤木材工業株式会社", "本社工場", 
         "愛知県春日井市新開町字平渕6", "0568-36-2931",
         "生産統括部", "部長 渡邉 茂芳", 20, "翌月20日",
         "8:00", "17:00", "20:00", "5:00",
         80, "7:40", 5, 1700, 1.25, 1.25, 1.35],
        ["F002", "加藤木材工業株式会社", "春日井工場",
         "愛知県春日井市新開町字別渕20", "0568-36-2932",
         "生産統括部", "課長 山田 太郎", 20, "翌月20日",
         "8:00", "17:00", "20:00", "5:00",
         80, "7:40", 5, 1700, 1.25, 1.25, 1.35],
        ["F003", "高雄工業株式会社", "岡山工場",
         "岡山県岡山市北区御津伊田1028-19", "086-724-5330",
         "岡山事業所", "課長 大治 国利", 15, "当月末日",
         "7:00", "15:30", "19:00", "3:30",
         45, "7:30", 15, 1650, 1.25, 1.25, 1.35],
    ]
    
    for row_idx, factory in enumerate(factories, 2):
        for col_idx, value in enumerate(factory, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 17:
                apply_style(cell, "currency")
            elif col_idx in [18, 19, 20]:
                apply_style(cell, "percent")
            else:
                apply_style(cell, "data")
    
    ws.freeze_panes = "A2"
    return ws


def create_kintai_sheet(wb, year, month):
    """Create 勤怠表 sheet for a specific month"""
    ws = wb.create_sheet(f"勤怠表_{year}年{month:02d}月")
    
    # Get days in month
    _, days_in_month = calendar.monthrange(year, month)
    
    # Header row 1 - Employee info
    info_headers = ["社員番号", "氏名", "派遣先", "工場名", "時給"]
    for col, header in enumerate(info_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_style(cell, "header")
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    # Header row 1 - Daily columns
    start_col = len(info_headers) + 1
    for day in range(1, days_in_month + 1):
        cell = ws.cell(row=1, column=start_col + (day-1)*4, value=f"{day}日")
        apply_style(cell, "header")
        ws.merge_cells(start_row=1, start_column=start_col + (day-1)*4, 
                       end_row=1, end_column=start_col + (day-1)*4 + 3)
    
    # Header row 2 - Sub columns for each day
    sub_headers = ["出勤", "退勤", "休憩", "実働"]
    for day in range(1, days_in_month + 1):
        for idx, sh in enumerate(sub_headers):
            col = start_col + (day-1)*4 + idx
            cell = ws.cell(row=2, column=col, value=sh)
            apply_style(cell, "subheader")
            ws.column_dimensions[get_column_letter(col)].width = 6
    
    # Summary columns at the end
    summary_start = start_col + days_in_month * 4
    summary_headers = [
        "出勤日数", "総労働時間", "所定内時間", "残業時間", 
        "深夜時間", "休日時間", "有休日数", "有休時間"
    ]
    for idx, header in enumerate(summary_headers):
        cell = ws.cell(row=1, column=summary_start + idx, value=header)
        apply_style(cell, "header")
        ws.merge_cells(start_row=1, start_column=summary_start + idx,
                       end_row=2, end_column=summary_start + idx)
        ws.column_dimensions[get_column_letter(summary_start + idx)].width = 10
    
    # Add sample employee data with formulas
    employees = [
        ("E001", "グエン・ヴァン・ミン", "加藤木材工業株式会社", "本社工場", 1700),
        ("E002", "チャン・ティ・ラン", "加藤木材工業株式会社", "本社工場", 1700),
        ("E003", "レ・ヴァン・ドゥック", "高雄工業株式会社", "岡山工場", 1650),
    ]
    
    for emp_idx, emp in enumerate(employees):
        row = emp_idx + 3
        for col_idx, value in enumerate(emp, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx == 5:
                apply_style(cell, "currency")
            else:
                apply_style(cell, "data")
        
        # Add sample time data for first 5 days
        for day in range(1, min(6, days_in_month + 1)):
            date = datetime(year, month, day)
            if date.weekday() < 5:  # Weekdays only
                col_base = start_col + (day-1)*4
                # 出勤 (start time)
                ws.cell(row=row, column=col_base, value="8:00")
                # 退勤 (end time)
                ws.cell(row=row, column=col_base + 1, value="17:00")
                # 休憩 (break minutes)
                ws.cell(row=row, column=col_base + 2, value=80)
                # 実働 (actual work hours) - Formula
                ws.cell(row=row, column=col_base + 3, 
                       value=f"=IF(OR({get_column_letter(col_base)}{row}=\"\",{get_column_letter(col_base+1)}{row}=\"\"),\"\","
                             f"({get_column_letter(col_base+1)}{row}-{get_column_letter(col_base)}{row})*24-{get_column_letter(col_base+2)}{row}/60)")
        
        # Summary formulas
        summary_col = summary_start
        # 出勤日数
        first_jissou_col = get_column_letter(start_col + 3)
        last_jissou_col = get_column_letter(start_col + (days_in_month-1)*4 + 3)
        ws.cell(row=row, column=summary_col, 
               value=f"=COUNTA({first_jissou_col}{row}:{last_jissou_col}{row})")
        # 総労働時間
        ws.cell(row=row, column=summary_col + 1,
               value=f"=SUMPRODUCT(({first_jissou_col}{row}:{last_jissou_col}{row}<>\"\")*(MOD(COLUMN({first_jissou_col}{row}:{last_jissou_col}{row})-COLUMN({first_jissou_col}{row}),4)=0)*({first_jissou_col}{row}:{last_jissou_col}{row}))")
    
    ws.freeze_panes = "F3"
    return ws


def create_salary_calculation(wb, year, month):
    """Create 給与計算 sheet"""
    ws = wb.create_sheet(f"給与計算_{year}年{month:02d}月")
    
    headers = [
        "社員番号", "氏名", "派遣先", "時給",
        "出勤日数", "所定内時間", "残業時間", "深夜時間", "休日時間", "有休時間",
        "基本給", "残業手当", "深夜手当", "休日手当", "有休手当", "交通費",
        "支給合計",
        "健康保険", "厚生年金", "雇用保険", "所得税", "住民税",
        "社宅家賃", "水道光熱費", "弁当代", "前貸",
        "控除合計", "差引支給額"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_style(cell, "header")
        ws.column_dimensions[get_column_letter(col)].width = 11
    
    # Add formulas for 3 sample employees
    employees = [
        ("E001", "グエン・ヴァン・ミン", "加藤木材工業株式会社", 1700),
        ("E002", "チャン・ティ・ラン", "加藤木材工業株式会社", 1700),
        ("E003", "レ・ヴァン・ドゥック", "高雄工業株式会社", 1650),
    ]
    
    for emp_idx, emp in enumerate(employees):
        row = emp_idx + 2
        
        # Basic info (columns A-D)
        for col, val in enumerate(emp, 1):
            cell = ws.cell(row=row, column=col, value=val)
            apply_style(cell, "currency" if col == 4 else "data")
        
        # Work hours - Sample data (E-J)
        sample_hours = [20, 160, 20, 10, 0, 8]  # days, regular, OT, night, holiday, paid leave
        for col, hours in enumerate(sample_hours, 5):
            cell = ws.cell(row=row, column=col, value=hours)
            apply_style(cell, "number")
        
        # Payment calculations (K-P)
        # 基本給 = 時給 × 所定内時間
        ws.cell(row=row, column=11, value=f"=D{row}*F{row}")
        # 残業手当 = 時給 × 残業時間 × 1.25
        ws.cell(row=row, column=12, value=f"=D{row}*G{row}*1.25")
        # 深夜手当 = 時給 × 深夜時間 × 0.25 (additional)
        ws.cell(row=row, column=13, value=f"=D{row}*H{row}*0.25")
        # 休日手当 = 時給 × 休日時間 × 1.35
        ws.cell(row=row, column=14, value=f"=D{row}*I{row}*1.35")
        # 有休手当 = 時給 × 有休時間
        ws.cell(row=row, column=15, value=f"=D{row}*J{row}")
        # 交通費
        ws.cell(row=row, column=16, value=5000)
        # 支給合計
        ws.cell(row=row, column=17, value=f"=SUM(K{row}:P{row})")
        
        for col in range(11, 18):
            apply_style(ws.cell(row=row, column=col), "currency")
        
        # Deductions (R-Z)
        # 健康保険 = 支給合計 × 5%
        ws.cell(row=row, column=18, value=f"=ROUND(Q{row}*0.05,0)")
        # 厚生年金 = 支給合計 × 9.15%
        ws.cell(row=row, column=19, value=f"=ROUND(Q{row}*0.0915,0)")
        # 雇用保険 = 支給合計 × 0.6%
        ws.cell(row=row, column=20, value=f"=ROUND(Q{row}*0.006,0)")
        # 所得税 (simplified - actual calculation is complex)
        ws.cell(row=row, column=21, value=f"=ROUND((Q{row}-88000)*0.05,0)")
        # 住民税 (fixed monthly amount)
        ws.cell(row=row, column=22, value=10000)
        # 社宅家賃
        ws.cell(row=row, column=23, value=25000)
        # 水道光熱費
        ws.cell(row=row, column=24, value=5000)
        # 弁当代
        ws.cell(row=row, column=25, value=f"=E{row}*400")
        # 前貸
        ws.cell(row=row, column=26, value=0)
        # 控除合計
        ws.cell(row=row, column=27, value=f"=SUM(R{row}:Z{row})")
        # 差引支給額
        ws.cell(row=row, column=28, value=f"=Q{row}-AA{row}")
        
        for col in range(18, 29):
            apply_style(ws.cell(row=row, column=col), "currency")
    
    # Total row
    total_row = len(employees) + 2
    ws.cell(row=total_row, column=1, value="合計")
    apply_style(ws.cell(row=total_row, column=1), "total")
    
    for col in [5, 6, 7, 8, 9, 10]:  # Hours totals
        ws.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row-1})")
        apply_style(ws.cell(row=total_row, column=col), "total")
    
    for col in range(11, 29):  # Money totals
        ws.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row-1})")
        apply_style(ws.cell(row=total_row, column=col), "total")
        ws.cell(row=total_row, column=col).number_format = "¥#,##0"
    
    ws.freeze_panes = "A2"
    return ws


def create_payslip(wb, year, month):
    """Create 給与明細 template sheet"""
    ws = wb.create_sheet(f"給与明細_{year}年{month:02d}月")
    
    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value=f"給 与 明 細 書  {year}年{month}月分")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Company info
    ws.cell(row=2, column=1, value=UNS_CONFIG["company_name"])
    ws.cell(row=3, column=1, value=f"派遣許可番号: {UNS_CONFIG['permit_number']}")
    
    # Employee info section
    sheet_ref = f"'給与計算_{year}年{month:02d}月'"
    ws.cell(row=5, column=1, value="社員番号")
    ws.cell(row=5, column=2, value=f"={sheet_ref}!A2")
    ws.cell(row=5, column=4, value="氏名")
    ws.cell(row=5, column=5, value=f"={sheet_ref}!B2")
    ws.cell(row=6, column=1, value="派遣先")
    ws.cell(row=6, column=2, value=f"={sheet_ref}!C2")
    
    # 勤怠欄
    ws.cell(row=8, column=1, value="【勤怠】")
    ws.cell(row=8, column=1).font = Font(bold=True)
    
    kintai_headers = ["出勤日数", "所定内時間", "残業時間", "深夜時間", "休日時間", "有休時間"]
    for idx, header in enumerate(kintai_headers):
        cell = ws.cell(row=9, column=idx+1, value=header)
        apply_style(cell, "subheader")
        ref_col = get_column_letter(5 + idx)
        ws.cell(row=10, column=idx+1, value=f"={sheet_ref}!{ref_col}2")
    
    # 支給欄
    ws.cell(row=12, column=1, value="【支給】")
    ws.cell(row=12, column=1).font = Font(bold=True)
    
    shikyuu_headers = ["基本給", "残業手当", "深夜手当", "休日手当", "有休手当", "交通費", "支給合計"]
    for idx, header in enumerate(shikyuu_headers):
        cell = ws.cell(row=13, column=idx+1, value=header)
        apply_style(cell, "subheader")
        ref_col = get_column_letter(11 + idx)
        cell = ws.cell(row=14, column=idx+1, value=f"={sheet_ref}!{ref_col}2")
        apply_style(cell, "currency")
    
    # 控除欄
    ws.cell(row=16, column=1, value="【控除】")
    ws.cell(row=16, column=1).font = Font(bold=True)
    
    koujo_headers = ["健康保険", "厚生年金", "雇用保険", "所得税", "住民税", "社宅家賃", "水道光熱費", "弁当代"]
    for idx, header in enumerate(koujo_headers):
        cell = ws.cell(row=17, column=idx+1, value=header)
        apply_style(cell, "subheader")
        ref_col = get_column_letter(18 + idx)
        cell = ws.cell(row=18, column=idx+1, value=f"={sheet_ref}!{ref_col}2")
        apply_style(cell, "currency")
    
    # 控除合計・差引支給額
    ws.cell(row=20, column=1, value="控除合計")
    ws.cell(row=20, column=1).font = Font(bold=True)
    cell = ws.cell(row=20, column=2, value=f"={sheet_ref}!AA2")
    apply_style(cell, "currency")
    
    ws.cell(row=22, column=1, value="差引支給額")
    ws.cell(row=22, column=1).font = Font(bold=True, size=14, color="FF0000")
    cell = ws.cell(row=22, column=2, value=f"={sheet_ref}!AB2")
    cell.font = Font(bold=True, size=14, color="FF0000")
    cell.number_format = "¥#,##0"
    
    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    return ws


def create_dashboard(wb):
    """Create ダッシュボード (summary) sheet"""
    ws = wb.active
    ws.title = "ダッシュボード"
    
    # Title
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1, value="UNS 勤怠管理システム")
    title.font = Font(bold=True, size=18)
    title.alignment = Alignment(horizontal="center")
    
    # System info
    ws.cell(row=3, column=1, value="会社名:")
    ws.cell(row=3, column=2, value=UNS_CONFIG["company_name"])
    ws.cell(row=4, column=1, value="許可番号:")
    ws.cell(row=4, column=2, value=UNS_CONFIG["permit_number"])
    ws.cell(row=5, column=1, value="作成日:")
    ws.cell(row=5, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
    
    # Sheet guide
    ws.cell(row=7, column=1, value="【シート構成】")
    ws.cell(row=7, column=1).font = Font(bold=True)
    
    sheets_info = [
        ("従業員マスタ", "派遣社員の基本情報を管理"),
        ("派遣先マスタ", "派遣先（工場）の情報・締め日・シフト等"),
        ("勤怠表_YYYY年MM月", "日々の出退勤記録"),
        ("給与計算_YYYY年MM月", "給与計算・控除計算"),
        ("給与明細_YYYY年MM月", "従業員への給与明細書"),
    ]
    
    for idx, (sheet_name, desc) in enumerate(sheets_info, 8):
        cell = ws.cell(row=idx, column=1, value=sheet_name)
        cell.font = Font(bold=True)
        ws.cell(row=idx, column=2, value=desc)
    
    # Configuration reference
    ws.cell(row=14, column=1, value="【控除率設定】")
    ws.cell(row=14, column=1).font = Font(bold=True)
    
    deductions = [
        ("健康保険", "5.0%"),
        ("厚生年金", "9.15%"),
        ("雇用保険", "0.6%"),
    ]
    for idx, (name, rate) in enumerate(deductions, 15):
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=rate)
    
    ws.cell(row=19, column=1, value="【割増率設定】")
    ws.cell(row=19, column=1).font = Font(bold=True)
    
    premiums = [
        ("時間外労働", "25%増 (×1.25)"),
        ("深夜労働", "25%増 (×1.25)"),
        ("休日労働", "35%増 (×1.35)"),
        ("60h超過", "50%増 (×1.50)"),
    ]
    for idx, (name, rate) in enumerate(premiums, 20):
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=rate)
    
    # Adjust widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    
    return ws


def main():
    """Generate complete UNS Kintai Management System workbook"""
    print("=" * 60)
    print("UNS 勤怠管理システム 生成中...")
    print("=" * 60)
    
    wb = Workbook()
    
    # Set target month (current or next month)
    today = datetime.today()
    year = today.year
    month = today.month
    
    print(f"\n対象期間: {year}年{month}月")
    
    # Create all sheets
    print("\n1. ダッシュボード作成中...")
    create_dashboard(wb)
    
    print("2. 従業員マスタ作成中...")
    create_employee_master(wb)
    
    print("3. 派遣先マスタ作成中...")
    create_factory_master(wb)
    
    print("4. 勤怠表作成中...")
    create_kintai_sheet(wb, year, month)
    
    print("5. 給与計算シート作成中...")
    create_salary_calculation(wb, year, month)
    
    print("6. 給与明細テンプレート作成中...")
    create_payslip(wb, year, month)
    
    # Save workbook
    output_file = f"UNS_勤怠システム_{year}{month:02d}.xlsx"
    wb.save(output_file)
    
    print("\n" + "=" * 60)
    print(f"✅ 完成: {output_file}")
    print("=" * 60)
    print("\n【シート一覧】")
    for sheet in wb.sheetnames:
        print(f"  • {sheet}")
    
    return output_file


if __name__ == "__main__":
    main()
