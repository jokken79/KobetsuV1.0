import openpyxl
import json
from openpyxl.utils import get_column_letter

def analyze_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    analysis = {
        "sheet_name": sheet.title,
        "dimensions": sheet.dimensions,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "merged_cells": [str(r) for r in sheet.merged_cells.ranges],
        "content_map": []
    }
    
    # Extract text content and their coordinates
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                analysis["content_map"].append({
                    "coord": cell.coordinate,
                    "value": str(cell.value).strip(),
                    "row": cell.row,
                    "col": cell.column,
                    "col_letter": get_column_letter(cell.column)
                })
                
    return analysis

if __name__ == "__main__":
    result = analyze_excel("outputs/final/01_個別契約書.xlsx")
    print(json.dumps(result, indent=2, ensure_ascii=False))
