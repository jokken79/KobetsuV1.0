import openpyxl
from openpyxl.utils import get_column_letter

def analyze_column_widths(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    print(f"Sheet: {sheet.title}")
    print("Column Widths:")
    
    total_width = 0
    widths = []
    
    # Check first 36 columns (A to AJ)
    for i in range(1, 37):
        col_letter = get_column_letter(i)
        # openpyxl reports width in characters (approx). 
        # Default is usually ~8.43.
        width = sheet.column_dimensions[col_letter].width
        if width is None:
            width = 8.43 # Default Excel width
        
        print(f"Col {col_letter}: {width}")
        widths.append(width)
        total_width += width
        
    print(f"\nTotal Width (Excel units): {total_width}")
    
    # Calculate percentage for each column
    print("\nPercentages:")
    for i, w in enumerate(widths):
        pct = (w / total_width) * 100
        print(f"Col {get_column_letter(i+1)}: {pct:.2f}%")

if __name__ == "__main__":
    analyze_column_widths("outputs/final/01_個別契約書.xlsx")
